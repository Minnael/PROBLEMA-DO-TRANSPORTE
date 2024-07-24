import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

class ConstrutorTabelas():
   def criar_tabela_excel(matriz, nome_arquivo):
      # Cria um novo workbook e seleciona a planilha ativa
      wb = Workbook()
      ws = wb.active

      # Define o preenchimento azul-claro para a primeira linha
      fill_azul_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

      # Define as bordas finas
      thin_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

      # Define a fonte Caveat, tamanho 14, negrito
      fonte = Font(name='Caveat', size=14, bold=True)

      # Preenche a planilha com a matriz
      for i, row in enumerate(matriz):
         for j, value in enumerate(row):
               cell = ws.cell(row=i + 1, column=j + 1, value=value)
               cell.border = thin_border
               cell.font = fonte
               cell.alignment = Alignment(horizontal='center', vertical='center')

               if i == 0:  # Se for a primeira linha
                  cell.fill = fill_azul_claro

      # Salva o arquivo
      wb.save(f"{nome_arquivo}.xlsx")
